import os
import openpyxl
import json

cd = os.getcwd()

def jsonify(sheet):
    headings = {}
    #print(sheet.max_row, sheet.max_column)

    for j in range(1, sheet.max_column+1):
        headings[j] = sheet.cell(row=1, column=j).value
        j += 1

    op = []
    for i in range(2, sheet.max_row+1):
        card = {}
        for j in range(1, sheet.max_column+1):
            val = sheet.cell(row=i, column=j).value
            if type(val) != str or val[0] != '#':
                card[headings[j]] = val
            else:
                print('Ignoring', i, j, headings[j], val)
        op.append(card)

    return op


def do_the_thing(xl):
    data = {}
    for key in xl.sheetnames:
        data[key.lower()] = jsonify(xl[key])
    return data


def main():
    xl = openpyxl.load_workbook(cd+'/data.xlsx', data_only=True)

    data = do_the_thing(xl)

    with open(cd+'/data.json', 'w') as f:
        json.dump(data, f)


if __name__ == '__main__':
    main()